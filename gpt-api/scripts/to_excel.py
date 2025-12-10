"""
JSON to Excel Converter
Converts lot boundary JSON to Excel format matching engineering closure tables.
"""

import json
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def load_json(json_path):
    """
    Load JSON data from file.
    
    Args:
        json_path: Path to JSON file
        
    Returns:
        Parsed JSON object
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data


def json_to_dataframe(data):
    """
    Convert JSON lot boundary data to pandas DataFrame.
    
    Args:
        data: JSON object with lots and boundaries
        
    Returns:
        pandas DataFrame with columns: Lot, Area, Distance, Bearing, Degrees, Minutes, Seconds
    """
    rows = []
    
    for lot in data["lots"]:
        lot_number = lot.get("lot", "")
        area = lot.get("area_m2", "")
        
        # Convert area to string, handle None/empty
        if area is None or area == "":
            area_str = ""
        else:
            area_str = str(area)
        
        boundaries = lot.get("boundaries", [])
        
        for i, boundary in enumerate(boundaries):
            row = {
                "Lot": lot_number if i == 0 else "",  # Only show lot number on first row
                "Area": area_str if i == 0 else "",   # Only show area on first row
                "Distance": boundary.get("length", ""),
                "Bearing": boundary.get("bearing", ""),
                "Degrees": boundary.get("deg", ""),
                "Minutes": boundary.get("min", ""),
                "Seconds": boundary.get("sec", "")
            }
            rows.append(row)
    
    df = pd.DataFrame(rows)
    return df


def save_excel(df, output_path):
    """
    Save DataFrame to Excel with proper formatting.
    
    Args:
        df: pandas DataFrame
        output_path: Path to save Excel file
    """
    # Create output directory if needed
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Write to Excel
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Lot Boundaries", index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets["Lot Boundaries"]
        
        # Format header row
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Center align numeric columns
        center_alignment = Alignment(horizontal="center", vertical="center")
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # Center Distance, Degrees, Minutes, Seconds
            for col_idx in [3, 5, 6, 7]:  # Distance, Degrees, Minutes, Seconds (1-indexed)
                if col_idx <= len(row):
                    row[col_idx - 1].alignment = center_alignment
    
    print(f"Excel file saved to: {output_path}")


def convert_json_to_excel(json_path, output_path=None):
    """
    Main function: Convert JSON to Excel.
    
    Args:
        json_path: Path to input JSON file
        output_path: Path to save Excel file (default: same name as JSON with .xlsx)
        
    Returns:
        Path to saved Excel file
    """
    json_path = Path(json_path)
    
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")
    
    if output_path is None:
        output_path = json_path.with_suffix(".xlsx")
    else:
        output_path = Path(output_path)
    
    # Load JSON
    print(f"Loading JSON from: {json_path}")
    data = load_json(str(json_path))
    
    # Convert to DataFrame
    df = json_to_dataframe(data)
    print(f"Converted {len(df)} boundary rows from {len(data['lots'])} lots")
    
    # Save to Excel
    save_excel(df, str(output_path))
    
    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python to_excel.py <json_path> [output_path]")
        sys.exit(1)
    
    json_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        excel_path = convert_json_to_excel(json_path, output_path)
        print(f"\n✅ Success! Excel saved to: {excel_path}")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        sys.exit(1)

